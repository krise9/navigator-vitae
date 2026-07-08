"""
Extract MBTI questions from the original Navigator Vitae Excel workbook.

Version: v2
Purpose:
- Reconstruct full question and answer text across multiple Excel rows.
- Preserve score letters, dimensions and weights from the Excel formulas.
- Generate questions/mbti.json for the Streamlit module.
- Generate data/debug_mbti_extraction.csv for manual inspection.

Expected project structure:
Navigator/
├── data/source/MBTI navigator vitae.xlsx
├── questions/
└── data/

Run from the Navigator project root:
python scripts/extract_mbti_from_excel.py
"""

from __future__ import annotations

import csv
import json
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

PROJECT_ROOT = Path(__file__).resolve().parents[1]
SOURCE_FILE = PROJECT_ROOT / "data" / "source" / "MBTI navigator vitae.xlsx"
OUTPUT_JSON = PROJECT_ROOT / "questions" / "mbti.json"
DEBUG_CSV = PROJECT_ROOT / "data" / "debug_mbti_extraction.csv"

SHEET_NAME = "Man"

LETTER_TO_DIMENSION = {
    "E": "EI", "I": "EI",
    "S": "SN", "N": "SN",
    "T": "TF", "F": "TF",
    "J": "JP", "P": "JP",
}

# Narrative sections: question number, question text, answer text, scoring formula column.
NARRATIVE_BLOCKS = [
    ("A", "B", "C", "AI"),
    ("O", "P", "Q", "AJ"),
    ("W", "X", "Y", "AK"),
]

# Word-pair sections: question number, word text, scoring formula column.
WORD_BLOCKS = [
    ("A", "C", "AI"),
    ("K", "M", "AJ"),
    ("S", "U", "AK"),
    ("AA", "AC", "AL"),
]

# Row ranges in the source workbook.
NARRATIVE_RANGES = [
    (4, 73, "deel_I"),
    (98, 155, "deel_III"),
]

WORD_RANGES = [
    (78, 94, "deel_II"),
    (172, 188, "deel_IV"),
]


def cell_value(ws: Worksheet, coord: str) -> str:
    value = ws[coord].value
    if value is None:
        return ""
    return str(value).strip()


def row_value(ws: Worksheet, col: str, row: int) -> str:
    return cell_value(ws, f"{col}{row}")


def parse_question_number(value: Any) -> Optional[int]:
    if value is None:
        return None
    text = str(value).strip()
    match = re.match(r"^(\d+)\.?$", text)
    if not match:
        return None
    return int(match.group(1))


def clean_piece(text: str) -> str:
    text = str(text).replace("\n", " ").strip()
    text = re.sub(r"\s+", " ", text)
    return text


def join_pieces(parts: List[str]) -> str:
    cleaned = []
    for part in parts:
        part = clean_piece(part)
        if not part:
            continue
        # In the source sheet, standalone "of" rows visually separate alternatives.
        if part.lower() == "of":
            continue
        cleaned.append(part)

    text = " ".join(cleaned)
    text = re.sub(r"\s+", " ", text)
    text = re.sub(r"\s+([,.;:?!])", r"\1", text)
    return text.strip()


def clean_answer(text: str, is_first_option: bool = False) -> str:
    text = join_pieces([text])
    if is_first_option:
        # Remove trailing connective that belongs to the visual layout, not the option.
        text = re.sub(r",?\s+of\s*$", "", text, flags=re.IGNORECASE).strip()
    text = re.sub(r"\s+([,.;:?!])", r"\1", text)
    return text.strip()


def parse_score_label(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip().replace(":", "").strip().upper()
    if text in LETTER_TO_DIMENSION:
        return text
    return None


def extract_weight(formula: Any) -> Optional[int]:
    """Extract the multiplier from formulas like '=B6*2'."""
    if formula is None:
        return None
    text = str(formula).strip()
    match = re.search(r"\*(\d+)", text)
    if not match:
        return None
    return int(match.group(1))


def has_formula_for_block(ws: Worksheet, score_col: str, row: int) -> bool:
    value = row_value(ws, score_col, row)
    return value.startswith("=")


def find_question_starts(ws: Worksheet, q_col: str, start_row: int, end_row: int) -> List[Tuple[int, int]]:
    starts = []
    for row in range(start_row, end_row + 1):
        number = parse_question_number(ws[f"{q_col}{row}"].value)
        if number is not None:
            starts.append((row, number))
    return starts


def find_label_rows(ws: Worksheet, score_col: str, start_row: int, end_row: int) -> List[Tuple[int, str, Optional[int]]]:
    labels = []
    for row in range(start_row, end_row):
        label = parse_score_label(ws[f"AH{row}"].value)
        if label and has_formula_for_block(ws, score_col, row):
            labels.append((row, label, extract_weight(ws[f"{score_col}{row}"].value)))
    return labels


def rows_text(ws: Worksheet, text_col: str, start_row: int, end_row: int) -> str:
    parts = []
    for row in range(start_row, end_row):
        value = row_value(ws, text_col, row)
        if value:
            parts.append(value)
    return join_pieces(parts)


def dimension_from_scores(a_score: str, b_score: str) -> str:
    a_dim = LETTER_TO_DIMENSION.get(a_score)
    b_dim = LETTER_TO_DIMENSION.get(b_score)
    if a_dim and a_dim == b_dim:
        return a_dim
    return "UNKNOWN"


def parse_narrative_question(
    ws: Worksheet,
    number: int,
    q_start: int,
    q_end: int,
    section: str,
    text_col: str,
    answer_col: str,
    score_col: str,
) -> Dict[str, Any]:
    label_rows = find_label_rows(ws, score_col, q_start, q_end)

    if len(label_rows) < 2:
        raise ValueError(
            f"Question {number}: expected at least 2 scoring label rows in {score_col}, found {label_rows}"
        )

    first_row, a_score, a_weight = label_rows[0]
    second_row, b_score, b_weight = label_rows[1]

    question_text = rows_text(ws, text_col, q_start, first_row)
    answer_a = clean_answer(rows_text(ws, answer_col, first_row, second_row), is_first_option=True)
    answer_b = clean_answer(rows_text(ws, answer_col, second_row, q_end), is_first_option=False)

    # Special note: the original workbook contains at least one item where the second option
    # spans a second scored row. This parser preserves the text in B, but Streamlit v0.1 still
    # treats the question as A/B.
    extra_scored_rows = []
    for row in range(second_row + 1, q_end):
        if has_formula_for_block(ws, score_col, row):
            extra_scored_rows.append(row)

    return {
        "id": f"MBTI{number:03d}",
        "excel_question_number": number,
        "section": section,
        "dimension": dimension_from_scores(a_score, b_score),
        "text": question_text,
        "A": answer_a,
        "B": answer_b,
        "A_score": a_score,
        "B_score": b_score,
        "A_weight": a_weight if a_weight is not None else 0,
        "B_weight": b_weight if b_weight is not None else 0,
        "source": {
            "question_rows": f"{q_start}:{q_end - 1}",
            "text_col": text_col,
            "answer_col": answer_col,
            "score_col": score_col,
            "a_row": first_row,
            "b_row": second_row,
            "extra_scored_rows": extra_scored_rows,
        },
    }


def parse_narrative_sections(ws: Worksheet) -> List[Dict[str, Any]]:
    questions = []

    for start_row, end_row, section in NARRATIVE_RANGES:
        for q_col, text_col, answer_col, score_col in NARRATIVE_BLOCKS:
            starts = find_question_starts(ws, q_col, start_row, end_row)
            for index, (q_start, number) in enumerate(starts):
                q_end = starts[index + 1][0] if index + 1 < len(starts) else end_row + 1
                questions.append(
                    parse_narrative_question(
                        ws=ws,
                        number=number,
                        q_start=q_start,
                        q_end=q_end,
                        section=section,
                        text_col=text_col,
                        answer_col=answer_col,
                        score_col=score_col,
                    )
                )

    return questions


def parse_word_sections(ws: Worksheet) -> List[Dict[str, Any]]:
    questions = []

    for start_row, end_row, section in WORD_RANGES:
        for q_col, word_col, score_col in WORD_BLOCKS:
            for row in range(start_row, end_row + 1):
                number = parse_question_number(ws[f"{q_col}{row}"].value)
                if number is None:
                    continue

                # The source workbook has a label error in Deel IV: question 88 is printed as 87.
                # If MBTI087 already exists later, the duplicate check will catch it.
                word_a = clean_answer(row_value(ws, word_col, row), is_first_option=False)
                word_b = clean_answer(row_value(ws, word_col, row + 1), is_first_option=False)
                a_score = parse_score_label(ws[f"AH{row}"].value)
                b_score = parse_score_label(ws[f"AH{row + 1}"].value)
                a_weight = extract_weight(ws[f"{score_col}{row}"].value)
                b_weight = extract_weight(ws[f"{score_col}{row + 1}"].value)

                if not (word_a and word_b and a_score and b_score):
                    continue

                questions.append(
                    {
                        "id": f"MBTI{number:03d}",
                        "excel_question_number": number,
                        "section": section,
                        "dimension": dimension_from_scores(a_score, b_score),
                        "text": "Welk van beide woorden spreekt u het meest aan?",
                        "A": word_a,
                        "B": word_b,
                        "A_score": a_score,
                        "B_score": b_score,
                        "A_weight": a_weight if a_weight is not None else 0,
                        "B_weight": b_weight if b_weight is not None else 0,
                        "source": {
                            "question_rows": f"{row}:{row + 1}",
                            "word_col": word_col,
                            "score_col": score_col,
                            "a_row": row,
                            "b_row": row + 1,
                        },
                    }
                )

    return questions


def fix_known_excel_label_issue(questions: List[Dict[str, Any]]) -> None:
    """
    Deel IV in the source contains a repeated label 87 where the sequence should be 88.
    We correct only the duplicate in section deel_IV with word pair Ordelijk/Ongedwongen.
    """
    seen = set()
    for question in questions:
        qid = question["id"]
        if qid in seen and question["section"] == "deel_IV" and question["A"] == "Ordelijk":
            question["excel_question_number"] = 88
            question["id"] = "MBTI088"
        seen.add(qid)


def validate_questions(questions: List[Dict[str, Any]]) -> Dict[str, Any]:
    ids = [q["id"] for q in questions]
    duplicates = sorted({qid for qid in ids if ids.count(qid) > 1})
    expected = {f"MBTI{i:03d}" for i in range(1, 95)}
    found = set(ids)
    missing = sorted(expected - found)
    extra = sorted(found - expected)
    unknown_dimensions = [q["id"] for q in questions if q["dimension"] == "UNKNOWN"]
    empty_text = [q["id"] for q in questions if not q["text"] or not q["A"] or not q["B"]]

    return {
        "duplicates": duplicates,
        "missing": missing,
        "extra": extra,
        "unknown_dimensions": unknown_dimensions,
        "empty_text": empty_text,
    }


def write_outputs(questions: List[Dict[str, Any]]) -> None:
    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    DEBUG_CSV.parent.mkdir(parents=True, exist_ok=True)

    # Sort by question number.
    questions.sort(key=lambda q: q["excel_question_number"])

    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(questions, f, ensure_ascii=False, indent=2)

    fieldnames = [
        "id", "section", "dimension", "text", "A", "B",
        "A_score", "B_score", "A_weight", "B_weight", "source",
    ]
    with open(DEBUG_CSV, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for q in questions:
            row = {key: q.get(key, "") for key in fieldnames}
            row["source"] = json.dumps(q.get("source", {}), ensure_ascii=False)
            writer.writerow(row)


def main() -> None:
    if not SOURCE_FILE.exists():
        raise FileNotFoundError(
            f"Excel source file not found: {SOURCE_FILE}\n"
            "Put the converted .xlsx file in data/source/."
        )

    wb = load_workbook(SOURCE_FILE, data_only=False)
    ws = wb[SHEET_NAME]

    questions = []
    questions.extend(parse_narrative_sections(ws))
    questions.extend(parse_word_sections(ws))
    fix_known_excel_label_issue(questions)

    validation = validate_questions(questions)
    write_outputs(questions)

    print(f"Extracted questions: {len(questions)}")
    print("Duplicates:", validation["duplicates"] or "none")
    print("Missing:", validation["missing"] or "none")
    print("Extra:", validation["extra"] or "none")
    print("Unknown dimensions:", validation["unknown_dimensions"] or "none")
    print("Empty text fields:", validation["empty_text"] or "none")
    print(f"Wrote: {OUTPUT_JSON.relative_to(PROJECT_ROOT)}")
    print(f"Wrote: {DEBUG_CSV.relative_to(PROJECT_ROOT)}")


if __name__ == "__main__":
    main()
